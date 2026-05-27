"""Importa as tabelas do `Precificador_Seguro_Automóvel_v2.0.xlsx` e gera
`src/insurmind/quote_tables.py` com dicts literais Python.

**Filosofia (decidida em 2026-05-26):** Excel não é fonte runtime; é spec
compilável. Rodar este script uma vez quando uma versão nova da planilha
chega, comitar o `quote_tables.py` gerado, e o runtime carrega só Python
puro (sem dep `openpyxl`, sem leitura de disco em prod).

**Uso:**
    .\\.venv\\Scripts\\python.exe scripts/import_precificador.py

Lê: `Precificador_Seguro_Automóvel_v2.0.xlsx` (raiz do repo)
Escreve: `src/insurmind/quote_tables.py`

**Inconsistências da planilha tratadas:** SKUs com FIPE = '-' em algum ano
viram entradas ausentes do índice (Dolphin Mini 2023, HB20 Comfort 2025,
HB20 Platinum 2024+2023, Kwid Iconic 2024+2023+2025). Em runtime,
`compute_quote` levanta `QuoteUnavailableError(modelo, ano, anos_disponiveis)`.
Detalhes em RELATORIO.md sessão "2026-05-26 — Integração do tarifador real".
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path
from textwrap import dedent

# Suprime warning chato de Data Validation (não afeta leitura)
warnings.filterwarnings("ignore", message="Data Validation extension")

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = REPO_ROOT / "Precificador_Seguro_Automóvel_v2.0.xlsx"
OUT = REPO_ROOT / "src" / "insurmind" / "quote_tables.py"


# Mapeamento ano-input → coluna FIPE na sheet MODELOS
ANO_TO_COL = {"0km": "H", "2026": "I", "2025": "J", "2024": "K", "2023": "L"}


def load() -> openpyxl.Workbook:
    return openpyxl.load_workbook(XLSX, data_only=True)


def extract_modelos(wb):
    """MODELOS: 16 SKUs com FIPE por ano + fator_modelo.

    Retorna:
        skus: list[dict] com {sku, chave, marca, modelo, versao, tipo, motor, fator_modelo, observacao}
        fipe_por_chave_ano: dict[(chave, ano)] → Decimal
    """
    ws = wb["MODELOS"]
    skus = []
    fipe = {}
    for row in range(2, 18):  # M01..M16 em linhas 2..17
        sku = ws.cell(row, 1).value
        if not sku:
            continue
        marca = ws.cell(row, 2).value
        modelo = ws.cell(row, 3).value
        versao = ws.cell(row, 4).value
        tipo = ws.cell(row, 5).value
        motor = ws.cell(row, 6).value
        # Cell G é fórmula =B&" "&C&" - "&D → openpyxl com data_only=True devolve o valor calculado
        chave = ws.cell(row, 7).value
        fator_modelo = ws.cell(row, 13).value
        observacao = ws.cell(row, 14).value or ""

        skus.append({
            "sku": sku,
            "chave": chave,
            "marca": marca,
            "modelo": modelo,
            "versao": versao,
            "tipo": tipo,
            "motor": motor,
            "fator_modelo": float(fator_modelo),
            "observacao": observacao,
        })

        # FIPE por ano: pula valores '-' (modelo não existe no ano)
        for ano, col in ANO_TO_COL.items():
            val = ws.cell(row, openpyxl.utils.column_index_from_string(col)).value
            if val is None or val == "-":
                continue
            fipe[(chave, ano)] = float(val)
    return skus, fipe


def extract_fatores(wb):
    """FATORES: 12 sub-tabelas. Retorna dict por categoria."""
    ws = wb["FATORES"]

    # Helper: lê um bloco rotulado começando em (row_label, col_chave, col_valor)
    def read_block(label_row, n_rows, col_chave=1, col_valor=2):
        result = {}
        for i in range(n_rows):
            r = label_row + i
            k = ws.cell(r, col_chave).value
            v = ws.cell(r, col_valor).value
            if k is None:
                continue
            result[str(k).strip()] = float(v) if isinstance(v, (int, float)) else v
        return result

    return {
        # Linhas exatas da planilha v2.0 conforme inspeção
        "taxa_casco":           float(ws["B5"].value),  # 0.03
        "taxa_rcfv":            float(ws["B6"].value),  # 0.008
        "taxa_app":             float(ws["B7"].value),  # 0.003
        "lmi_rcfv":             float(ws["B12"].value),  # 100000
        "lmi_app":              float(ws["B13"].value),  # 20000
        "fator_cobertura_casco": read_block(17, 3),       # Compreensiva/RF+Inc+RCF-V/Só RCF-V
        "fator_cobertura_rcfv":  read_block(24, 3),
        "fator_cobertura_app":   read_block(31, 3),
        "fator_franquia":        read_block(38, 3),       # Reduzida/Normal/Aumentada
        "fator_idade":           read_block(45, 6),       # 18-25/26-30/31-40/41-55/56-65/66+
        "fator_sexo":            read_block(55, 2),       # Masculino/Feminino
        "fator_uso":             read_block(61, 4),       # 4 categorias
        "fator_garagem":         read_block(69, 3),       # 3 categorias
        "fator_bonus":           read_block(76, 11),      # Classe 0..Classe 10
        "valor_assistencia":     read_block(91, 2),       # Básica/Ampliada (R$)
        "carregamento":         float(ws["B97"].value),  # 0.35
        "iof":                  float(ws["B98"].value),  # 0.0738
    }


def extract_capitais(wb):
    """CAPITAIS: 6 capitais → fator região."""
    ws = wb["CAPITAIS"]
    capitais = {}
    for row in range(3, 9):  # linhas 3..8
        nome = ws.cell(row, 1).value
        if not nome:
            continue
        capitais[nome] = float(ws.cell(row, 3).value)
    return capitais


def render(skus, fipe, fatores, capitais):
    """Gera o conteúdo Python literal do quote_tables.py."""
    from decimal import Decimal

    header = dedent('''\
        """Tabelas do tarifador, geradas a partir de `Precificador_Seguro_Automóvel_v2.0.xlsx`.

        **NÃO EDITAR À MÃO.** Regerar com `python scripts/import_precificador.py`
        quando o grupo (João Carlos + Adriele) entregar uma versão nova da planilha.

        Fonte: planilha v2.0 (modificada em 2026-05-22), criada por João Carlos Mendonça.
        Importada em: gerada automaticamente — ver topo do arquivo gerado.

        Inconsistências da planilha tratadas:
        - SKUs com FIPE = '-' em algum ano são omitidas do índice
          `FIPE_POR_CHAVE_ANO`. Em runtime, levanta QuoteUnavailableError.
        - Fórmula da LEIA-ME (A13) consolida F_Idade/F_Uso/F_Bônus aplicados
          ao prêmio puro total; as CÉLULAS aplicam por componente. As células
          são a verdade canônica — replicamos exatamente, inclusive a não-aplicação
          de F_Sexo e F_Garagem no APP (atuarialmente: APP indeniza passageiros,
          não depende do condutor).
        """
        from __future__ import annotations

        from decimal import Decimal


    ''')

    def fmt_dec(v):
        return f'Decimal("{v}")'

    # MODELOS
    lines = [header, "# 16 SKUs (8 modelos × 2 versões) — cada um com FIPE por ano + fator_modelo.\n"]
    lines.append("MODELOS: list[dict] = [\n")
    for s in skus:
        lines.append(
            "    {"
            f'"sku": {s["sku"]!r}, '
            f'"chave": {s["chave"]!r}, '
            f'"marca": {s["marca"]!r}, '
            f'"modelo": {s["modelo"]!r}, '
            f'"versao": {s["versao"]!r}, '
            f'"tipo": {s["tipo"]!r}, '
            f'"motor": {s["motor"]!r}, '
            f'"fator_modelo": {fmt_dec(s["fator_modelo"])}, '
            f'"observacao": {s["observacao"]!r}'
            "},\n"
        )
    lines.append("]\n\n")

    # Chave → fator_modelo (acesso rápido)
    lines.append("FATOR_MODELO_POR_CHAVE: dict[str, Decimal] = {\n")
    for s in skus:
        lines.append(f"    {s['chave']!r}: {fmt_dec(s['fator_modelo'])},\n")
    lines.append("}\n\n")

    # FIPE por (chave, ano)
    lines.append(
        "# FIPE por (chave, ano). Pares ausentes = SKU não existe naquele ano (foi '-' na planilha).\n"
    )
    lines.append("FIPE_POR_CHAVE_ANO: dict[tuple[str, str], Decimal] = {\n")
    for (chave, ano), v in fipe.items():
        lines.append(f"    ({chave!r}, {ano!r}): {fmt_dec(v)},\n")
    lines.append("}\n\n")

    # FATORES
    lines.append("# Taxas-base por componente (sobre IS_FIPE / LMI_RCFV / LMI_APP).\n")
    lines.append(f"TAXA_CASCO = {fmt_dec(fatores['taxa_casco'])}\n")
    lines.append(f"TAXA_RCFV  = {fmt_dec(fatores['taxa_rcfv'])}\n")
    lines.append(f"TAXA_APP   = {fmt_dec(fatores['taxa_app'])}\n\n")
    lines.append(f"LMI_RCFV   = {fmt_dec(fatores['lmi_rcfv'])}\n")
    lines.append(f"LMI_APP    = {fmt_dec(fatores['lmi_app'])}\n\n")

    def render_factor_dict(name, d):
        lines.append(f"{name}: dict[str, Decimal] = {{\n")
        for k, v in d.items():
            lines.append(f"    {k!r}: {fmt_dec(v)},\n")
        lines.append("}\n\n")

    render_factor_dict("FATOR_COBERTURA_CASCO", fatores["fator_cobertura_casco"])
    render_factor_dict("FATOR_COBERTURA_RCFV", fatores["fator_cobertura_rcfv"])
    render_factor_dict("FATOR_COBERTURA_APP", fatores["fator_cobertura_app"])
    render_factor_dict("FATOR_FRANQUIA", fatores["fator_franquia"])
    render_factor_dict("FATOR_IDADE", fatores["fator_idade"])
    render_factor_dict("FATOR_SEXO", fatores["fator_sexo"])
    render_factor_dict("FATOR_USO", fatores["fator_uso"])
    render_factor_dict("FATOR_GARAGEM", fatores["fator_garagem"])
    render_factor_dict("FATOR_BONUS", fatores["fator_bonus"])
    render_factor_dict("VALOR_ASSISTENCIA", fatores["valor_assistencia"])

    lines.append(f"CARREGAMENTO = {fmt_dec(fatores['carregamento'])}  # 35%: corretagem + admin + margem\n")
    lines.append(f"IOF          = {fmt_dec(fatores['iof'])}  # 7.38% sobre o prêmio líquido\n\n")

    # CAPITAIS
    render_factor_dict("FATOR_REGIAO", capitais)

    return "".join(lines)


def main():
    print(f"Lendo: {XLSX.name}")
    wb = load()
    skus, fipe = extract_modelos(wb)
    fatores = extract_fatores(wb)
    capitais = extract_capitais(wb)

    print(f"  SKUs: {len(skus)}")
    print(f"  FIPE entries: {len(fipe)} (esperado ≤ 80 = 16×5; ausências = SKU sem FIPE no ano)")
    print(f"  Capitais: {len(capitais)}")
    print(f"  Faixas etárias: {list(fatores['fator_idade'].keys())}")
    print(f"  Coberturas: {list(fatores['fator_cobertura_casco'].keys())}")
    print(f"  Franquias: {list(fatores['fator_franquia'].keys())}")
    print(f"  Classes bônus: {len(fatores['fator_bonus'])} (esperado 11)")

    content = render(skus, fipe, fatores, capitais)
    OUT.write_text(content, encoding="utf-8")
    print(f"\nEscrito: {OUT.relative_to(REPO_ROOT)} ({len(content):,} chars)")


if __name__ == "__main__":
    main()
